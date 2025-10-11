"""
Comprehensive Sample Data Generator for Product Reviews
Creates a realistic dataset based on the provided schema for sentiment analysis
"""

import pandas as pd
import numpy as np
import random
import json
from datetime import datetime, timedelta
from typing import List, Dict
import uuid


def create_comprehensive_review_dataset(num_reviews: int = 2000) -> pd.DataFrame:
    """
    Create a comprehensive product review dataset based on the provided schema
    
    Args:
        num_reviews (int): Number of reviews to generate
        
    Returns:
        pd.DataFrame: Generated review dataset
    """
    np.random.seed(42)
    random.seed(42)
    
    # Product Information
    products = generate_product_catalog()
    
    # Customer Information
    customers = generate_customer_base()
    
    # Review templates
    review_templates = generate_review_templates()
    
    reviews = []
    
    for i in range(num_reviews):
        # Select random product and customer
        product = random.choice(products)
        customer = random.choice(customers)
        
        # Generate rating with realistic distribution
        rating = np.random.choice([1, 2, 3, 4, 5], p=[0.05, 0.1, 0.15, 0.35, 0.35])
        
        # Generate review content
        review_content = generate_review_content(review_templates, rating, product)
        
        # Generate engagement metrics
        helpful_votes = np.random.poisson(3) if rating >= 4 else np.random.poisson(1)
        comments_count = np.random.poisson(1) if helpful_votes > 5 else 0
        
        # Generate review date (last 2 years)
        start_date = datetime.now() - timedelta(days=730)
        random_days = random.randint(0, 730)
        review_date = start_date + timedelta(days=random_days)
        
        # Generate sentiment label based on rating
        if rating >= 4:
            sentiment_label = "Positive"
        elif rating == 3:
            sentiment_label = "Neutral"
        else:
            sentiment_label = "Negative"
        
        # Generate aspect sentiments
        aspect_sentiments = generate_aspect_sentiments(product, rating)
        
        # Generate media URLs (optional)
        media_urls = generate_media_urls() if random.random() > 0.8 else ""
        
        review = {
            # Review/Feedback Data
            'Review_ID': f"REV_{str(i).zfill(6)}",
            'Review_Text': review_content['text'],
            'Review_Title': review_content['title'],
            'Rating': rating,
            'Review_Date': review_date.strftime('%Y-%m-%d'),
            'Verified_Purchase': random.choice([True, False]),
            
            # Product Information
            'Product_ID': product['Product_ID'],
            'Product_Name': product['Product_Name'],
            'Category': product['Category'],
            'Subcategory': product['Subcategory'],
            'Brand': product['Brand'],
            'Price': product['Price'],
            'Features': product['Features'],
            
            # Customer Information
            'Customer_ID': customer['Customer_ID'],
            'Customer_Name': customer['Customer_Name'],
            'Age_Group': customer['Age_Group'],
            'Location': customer['Location'],
            'Purchase_History': customer['Purchase_History'],
            
            # Engagement Metrics
            'Helpful_Votes': helpful_votes,
            'Comments_Count': comments_count,
            'Media_URLs': media_urls,
            
            # Sentiment/Labels
            'Sentiment_Label': sentiment_label,
            'Aspect_Sentiments': json.dumps(aspect_sentiments)
        }
        
        reviews.append(review)
    
    df = pd.DataFrame(reviews)
    
    # Print summary
    print(f"Generated {len(df)} comprehensive product reviews")
    print(f"\nRating distribution:")
    rating_counts = df['Rating'].value_counts().sort_index()
    for rating, count in rating_counts.items():
        print(f"  {rating} stars: {count} reviews ({count/len(df)*100:.1f}%)")
    
    print(f"\nSentiment distribution:")
    sentiment_counts = df['Sentiment_Label'].value_counts()
    for sentiment, count in sentiment_counts.items():
        print(f"  {sentiment}: {count} reviews ({count/len(df)*100:.1f}%)")
    
    print(f"\nCategory distribution:")
    category_counts = df['Category'].value_counts()
    for category, count in category_counts.head(5).items():
        print(f"  {category}: {count} reviews")
    
    return df


def generate_product_catalog() -> List[Dict]:
    """Generate a comprehensive product catalog"""
    
    products = [
        # Electronics
        {"Product_ID": "PROD_001", "Product_Name": "iPhone 15 Pro Max", "Category": "Electronics", 
         "Subcategory": "Smartphones", "Brand": "Apple", "Price": 1199.99, 
         "Features": "A17 Pro chip, 48MP camera, Titanium design, USB-C"},
        
        {"Product_ID": "PROD_002", "Product_Name": "Samsung Galaxy S24 Ultra", "Category": "Electronics", 
         "Subcategory": "Smartphones", "Brand": "Samsung", "Price": 1299.99, 
         "Features": "Snapdragon 8 Gen 3, 200MP camera, S Pen, AI features"},
        
        {"Product_ID": "PROD_003", "Product_Name": "MacBook Pro 16-inch", "Category": "Electronics", 
         "Subcategory": "Laptops", "Brand": "Apple", "Price": 2499.99, 
         "Features": "M3 Pro chip, Liquid Retina XDR, 22-hour battery"},
        
        {"Product_ID": "PROD_004", "Product_Name": "Dell XPS 15", "Category": "Electronics", 
         "Subcategory": "Laptops", "Brand": "Dell", "Price": 1899.99, 
         "Features": "Intel i7, 4K OLED display, InfinityEdge design"},
        
        {"Product_ID": "PROD_005", "Product_Name": "Sony WH-1000XM5", "Category": "Electronics", 
         "Subcategory": "Headphones", "Brand": "Sony", "Price": 399.99, 
         "Features": "Noise canceling, 30-hour battery, LDAC codec"},
        
        {"Product_ID": "PROD_006", "Product_Name": "AirPods Pro 2nd Gen", "Category": "Electronics", 
         "Subcategory": "Earbuds", "Brand": "Apple", "Price": 249.99, 
         "Features": "Active noise canceling, Spatial Audio, H2 chip"},
        
        # Fashion
        {"Product_ID": "PROD_007", "Product_Name": "Nike Air Max 270", "Category": "Fashion", 
         "Subcategory": "Sneakers", "Brand": "Nike", "Price": 150.00, 
         "Features": "Air Max sole, Breathable mesh, Casual style"},
        
        {"Product_ID": "PROD_008", "Product_Name": "Levi's 501 Original Jeans", "Category": "Fashion", 
         "Subcategory": "Jeans", "Brand": "Levi's", "Price": 89.99, 
         "Features": "Straight fit, 100% cotton, Classic style"},
        
        {"Product_ID": "PROD_009", "Product_Name": "Adidas Ultraboost 22", "Category": "Fashion", 
         "Subcategory": "Running Shoes", "Brand": "Adidas", "Price": 180.00, 
         "Features": "Boost midsole, Primeknit upper, Continental rubber"},
        
        # Home & Garden
        {"Product_ID": "PROD_010", "Product_Name": "Instant Pot Duo 7-in-1", "Category": "Home & Garden", 
         "Subcategory": "Kitchen Appliances", "Brand": "Instant Pot", "Price": 99.99, 
         "Features": "Pressure cooker, Slow cooker, Rice cooker, Steamer"},
        
        {"Product_ID": "PROD_011", "Product_Name": "Dyson V15 Detect", "Category": "Home & Garden", 
         "Subcategory": "Vacuum Cleaners", "Brand": "Dyson", "Price": 749.99, 
         "Features": "Laser dust detection, 60-minute runtime, HEPA filtration"},
        
        {"Product_ID": "PROD_012", "Product_Name": "KitchenAid Stand Mixer", "Category": "Home & Garden", 
         "Subcategory": "Kitchen Appliances", "Brand": "KitchenAid", "Price": 329.99, 
         "Features": "5-quart bowl, 10 speeds, Multiple attachments"},
        
        # Books
        {"Product_ID": "PROD_013", "Product_Name": "The Seven Husbands of Evelyn Hugo", "Category": "Books", 
         "Subcategory": "Fiction", "Brand": "Simon & Schuster", "Price": 16.99, 
         "Features": "Paperback, 400 pages, Bestseller"},
        
        {"Product_ID": "PROD_014", "Product_Name": "Atomic Habits", "Category": "Books", 
         "Subcategory": "Self-Help", "Brand": "Avery", "Price": 18.99, 
         "Features": "Hardcover, 320 pages, Personal development"},
        
        # Beauty & Personal Care
        {"Product_ID": "PROD_015", "Product_Name": "Olaplex No.3 Hair Perfector", "Category": "Beauty & Personal Care", 
         "Subcategory": "Hair Care", "Brand": "Olaplex", "Price": 28.00, 
         "Features": "Repair treatment, Bond building, Professional formula"},
        
        {"Product_ID": "PROD_016", "Product_Name": "La Mer Crème de la Mer", "Category": "Beauty & Personal Care", 
         "Subcategory": "Skincare", "Brand": "La Mer", "Price": 195.00, 
         "Features": "Moisturizing cream, Sea kelp extract, Luxury skincare"},
        
        # Sports & Outdoors
        {"Product_ID": "PROD_017", "Product_Name": "Peloton Bike+", "Category": "Sports & Outdoors", 
         "Subcategory": "Exercise Equipment", "Brand": "Peloton", "Price": 2495.00, 
         "Features": "Interactive classes, 24-inch touchscreen, Auto-follow resistance"},
        
        {"Product_ID": "PROD_018", "Product_Name": "Yeti Rambler 30oz Tumbler", "Category": "Sports & Outdoors", 
         "Subcategory": "Drinkware", "Brand": "Yeti", "Price": 35.00, 
         "Features": "Double-wall insulation, BPA-free, Dishwasher safe"},
        
        # Toys & Games
        {"Product_ID": "PROD_019", "Product_Name": "LEGO Creator Expert Assembly Square", "Category": "Toys & Games", 
         "Subcategory": "Building Sets", "Brand": "LEGO", "Price": 279.99, 
         "Features": "4002 pieces, Modular building, Collectible"},
        
        {"Product_ID": "PROD_020", "Product_Name": "PlayStation 5 Console", "Category": "Toys & Games", 
         "Subcategory": "Gaming Consoles", "Brand": "Sony", "Price": 499.99, 
         "Features": "4K gaming, Ray tracing, SSD storage, DualSense controller"}
    ]
    
    return products


def generate_customer_base() -> List[Dict]:
    """Generate a diverse customer base"""
    
    customers = []
    
    # Generate 100 unique customers
    for i in range(100):
        age_groups = ["18-25", "26-35", "36-45", "46-55", "56-65", "65+"]
        locations = ["New York", "Los Angeles", "Chicago", "Houston", "Phoenix", "Philadelphia", 
                    "San Antonio", "San Diego", "Dallas", "San Jose", "Austin", "Jacksonville",
                    "Fort Worth", "Columbus", "Charlotte", "San Francisco", "Indianapolis",
                    "Seattle", "Denver", "Washington"]
        
        purchase_history = random.choice(["Frequent", "Regular", "Occasional", "New Customer"])
        
        customer = {
            "Customer_ID": f"CUST_{str(i).zfill(4)}",
            "Customer_Name": f"Customer_{i+1}",
            "Age_Group": random.choice(age_groups),
            "Location": random.choice(locations),
            "Purchase_History": purchase_history
        }
        
        customers.append(customer)
    
    return customers


def generate_review_templates() -> Dict:
    """Generate comprehensive review templates"""
    
    return {
        "positive": {
            "titles": [
                "Amazing product!",
                "Exceeded my expectations",
                "Perfect purchase",
                "Highly recommend",
                "Great quality",
                "Love it!",
                "Excellent value",
                "Outstanding product",
                "Fantastic quality",
                "Best purchase ever"
            ],
            "texts": [
                "This product is absolutely fantastic! The quality is outstanding and it works exactly as described. I've been using it for a while now and couldn't be happier with my purchase. The build quality is excellent and it's definitely worth the price. I would highly recommend this to anyone looking for a reliable product.",
                
                "I'm so impressed with this purchase! The product arrived quickly and in perfect condition. The quality is top-notch and it has exceeded all my expectations. The features work flawlessly and the design is beautiful. This is definitely one of my best purchases this year.",
                
                "Outstanding product! I've been looking for something like this for a while and this one is perfect. The quality is excellent, the price is fair, and the customer service was great. I would definitely buy from this brand again. Highly recommended!",
                
                "This is exactly what I needed! The product works perfectly and the quality is amazing. I've been using it daily and it has made my life so much easier. The design is sleek and modern, and it's built to last. Worth every penny!",
                
                "Fantastic product! I'm so glad I chose this one. The quality is outstanding and it works better than I expected. The packaging was perfect and it arrived on time. I would definitely recommend this to friends and family. Great purchase!"
            ]
        },
        "neutral": {
            "titles": [
                "Decent product",
                "Okay for the price",
                "Average quality",
                "It's fine",
                "Meets expectations",
                "Standard product",
                "Nothing special",
                "Fair quality",
                "Acceptable",
                "It works"
            ],
            "texts": [
                "This product is okay. It works as expected and the quality is decent for the price point. Nothing particularly outstanding about it, but it gets the job done. The design is simple and functional. I would consider it a fair purchase.",
                
                "The product is fine. It does what it's supposed to do without any major issues. The quality is average and the price seems reasonable. It's not the best product I've ever used, but it's not the worst either. It's a solid, middle-of-the-road option.",
                
                "This is a standard product that works adequately. The quality is acceptable and it serves its purpose well. There's nothing particularly impressive about it, but there's also nothing wrong with it. It's a reliable, no-frills option.",
                
                "The product meets my basic expectations. It's functional and does what it needs to do. The quality is fair for the price, though it's not exceptional. It's a practical choice that gets the job done without any surprises.",
                
                "This is an average product that works fine. The quality is decent and the price is reasonable. It's not going to blow your mind, but it's reliable and functional. It's a good choice if you're looking for something straightforward."
            ]
        },
        "negative": {
            "titles": [
                "Disappointed",
                "Poor quality",
                "Not worth it",
                "Waste of money",
                "Terrible product",
                "Don't recommend",
                "Very disappointed",
                "Poor construction",
                "Not as described",
                "Regret buying"
            ],
            "texts": [
                "I'm really disappointed with this product. The quality is poor and it doesn't work as advertised. It broke after just a few uses and the customer service was unhelpful. I would not recommend this to anyone. Save your money and look elsewhere.",
                
                "This product is terrible! The quality is awful and it's definitely not worth the price. It arrived damaged and doesn't function properly. The materials feel cheap and it's poorly constructed. I regret this purchase completely.",
                
                "Very disappointed with this purchase. The product doesn't match the description at all. The quality is subpar and it stopped working after a short time. The packaging was also damaged upon arrival. I would not buy this again.",
                
                "This is a waste of money. The product is poorly made and doesn't work as expected. The quality is terrible and it's not worth the price. I've had much better products for less money. I would not recommend this to anyone.",
                
                "Poor product quality and terrible customer service. The item arrived broken and when I tried to return it, the process was a nightmare. The product itself is cheaply made and doesn't last. Avoid this purchase at all costs."
            ]
        }
    }


def generate_review_content(templates: Dict, rating: int, product: Dict) -> Dict:
    """Generate review content based on rating and product"""
    
    if rating >= 4:
        sentiment = "positive"
    elif rating == 3:
        sentiment = "neutral"
    else:
        sentiment = "negative"
    
    title = random.choice(templates[sentiment]["titles"])
    text = random.choice(templates[sentiment]["texts"])
    
    # Add product-specific details
    product_mention = f" The {product['Product_Name']} is "
    if sentiment == "positive":
        text = text.replace("This product", f"The {product['Product_Name']}")
    elif sentiment == "negative":
        text = text.replace("This product", f"The {product['Product_Name']}")
    
    return {"title": title, "text": text}


def generate_aspect_sentiments(product: Dict, rating: int) -> Dict:
    """Generate aspect-based sentiments"""
    
    aspects = {
        "Electronics": ["performance", "design", "battery", "camera", "display"],
        "Fashion": ["fit", "style", "comfort", "durability", "material"],
        "Home & Garden": ["functionality", "design", "ease_of_use", "durability", "value"],
        "Books": ["content", "writing", "plot", "characters", "value"],
        "Beauty & Personal Care": ["effectiveness", "texture", "scent", "packaging", "value"],
        "Sports & Outdoors": ["performance", "comfort", "durability", "design", "value"],
        "Toys & Games": ["fun_factor", "quality", "educational_value", "durability", "value"]
    }
    
    product_aspects = aspects.get(product["Category"], ["quality", "design", "value", "functionality"])
    aspect_sentiments = {}
    
    for aspect in product_aspects:
        if rating >= 4:
            aspect_sentiments[aspect] = "positive"
        elif rating == 3:
            aspect_sentiments[aspect] = "neutral"
        else:
            aspect_sentiments[aspect] = "negative"
    
    return aspect_sentiments


def generate_media_urls() -> str:
    """Generate sample media URLs"""
    
    sample_urls = [
        "https://example.com/review_images/photo1.jpg",
        "https://example.com/review_images/photo2.jpg",
        "https://example.com/review_videos/video1.mp4"
    ]
    
    # Randomly select 0-2 URLs
    num_urls = random.randint(0, 2)
    selected_urls = random.sample(sample_urls, num_urls)
    
    return "; ".join(selected_urls)


if __name__ == "__main__":
    # Create comprehensive sample dataset
    print("Generating comprehensive product review dataset...")
    sample_data = create_comprehensive_review_dataset(2000)
    
    # Save to CSV
    sample_data.to_csv('data/comprehensive_reviews.csv', index=False)
    print(f"\nDataset saved to 'data/comprehensive_reviews.csv'")
    
    # Save to Excel with multiple sheets
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main data sheet
        ws_data = wb.create_sheet("Product Reviews Data")
        
        # Add data to worksheet
        for r in dataframe_to_rows(sample_data, index=False, header=True):
            ws_data.append(r)
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws_data[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Create dataset summary sheet
        ws_summary = wb.create_sheet("Dataset Summary")
        
        # Add summary information
        summary_data = [
            ["Dataset Information", ""],
            ["Total Reviews", len(sample_data)],
            ["Total Columns", len(sample_data.columns)],
            ["Date Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Rating Distribution", ""],
        ]
        
        # Add rating distribution
        rating_counts = sample_data['Rating'].value_counts().sort_index()
        for rating, count in rating_counts.items():
            summary_data.append([f"{rating} Stars", f"{count} ({count/len(sample_data)*100:.1f}%)"])
        
        summary_data.extend([
            ["", ""],
            ["Sentiment Distribution", ""],
        ])
        
        # Add sentiment distribution
        sentiment_counts = sample_data['Sentiment_Label'].value_counts()
        for sentiment, count in sentiment_counts.items():
            summary_data.append([sentiment, f"{count} ({count/len(sample_data)*100:.1f}%)"])
        
        summary_data.extend([
            ["", ""],
            ["Category Distribution", ""],
        ])
        
        # Add category distribution
        category_counts = sample_data['Category'].value_counts()
        for category, count in category_counts.items():
            summary_data.append([category, f"{count} reviews"])
        
        # Add summary data to worksheet
        for row in summary_data:
            ws_summary.append(row)
        
        # Style summary sheet
        for cell in ws_summary[1]:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Create column descriptions sheet
        ws_columns = wb.create_sheet("Column Descriptions")
        
        column_descriptions = [
            ["Column Name", "Data Type", "Description", "Required"],
            ["Review_ID", "String", "Unique identifier for the review", "Yes"],
            ["Review_Text", "Text", "Full review content (main input for sentiment analysis)", "Yes"],
            ["Review_Title", "String", "Optional short title or summary", "No"],
            ["Rating", "Integer", "Numeric rating (1–5)", "Yes"],
            ["Review_Date", "Date", "Date of review submission", "No"],
            ["Verified_Purchase", "Boolean", "True if review is from verified purchase", "No"],
            ["Product_ID", "String", "Unique identifier for the product", "No"],
            ["Product_Name", "String", "Name/title of the product", "No"],
            ["Category", "String", "Main product category (e.g., Electronics, Fashion)", "No"],
            ["Subcategory", "String", "More specific category (e.g., Smartphones, Shoes)", "No"],
            ["Brand", "String", "Brand/manufacturer", "No"],
            ["Price", "Float", "Product price", "No"],
            ["Features", "String", "Key product features (comma-separated or JSON format)", "No"],
            ["Customer_ID", "String", "Unique identifier for the customer", "No"],
            ["Customer_Name", "String", "Customer name", "No"],
            ["Age_Group", "String", "Age group demographic", "No"],
            ["Location", "String", "Customer location", "No"],
            ["Purchase_History", "String", "Customer purchase history type", "No"],
            ["Helpful_Votes", "Integer", "Number of helpful votes", "No"],
            ["Comments_Count", "Integer", "Number of replies/comments", "No"],
            ["Media_URLs", "Text", "Links to images/videos uploaded by reviewer", "No"],
            ["Sentiment_Label", "String", "Pre-labeled sentiment (Positive/Negative/Neutral)", "No"],
            ["Aspect_Sentiments", "JSON", "Feature-based sentiment analysis", "No"]
        ]
        
        # Add column descriptions
        for row in column_descriptions:
            ws_columns.append(row)
        
        # Style column descriptions sheet
        for cell in ws_columns[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths for descriptions
        for column in ws_columns.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_columns.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel file
        excel_filename = 'data/comprehensive_reviews_dataset.xlsx'
        wb.save(excel_filename)
        print(f"\nExcel dataset saved to '{excel_filename}'")
        print("Excel file contains:")
        print("- Product Reviews Data: Main dataset with all reviews")
        print("- Dataset Summary: Statistics and distributions")
        print("- Column Descriptions: Detailed column information")
        
    except ImportError:
        print("\nNote: openpyxl not installed. Install with: pip install openpyxl")
        print("Saving as CSV only...")
    
    # Show sample
    print("\nSample data preview:")
    print(sample_data.head())
    
    print("\nDataset columns:")
    print(sample_data.columns.tolist())
    
    print(f"\nDataset shape: {sample_data.shape}")